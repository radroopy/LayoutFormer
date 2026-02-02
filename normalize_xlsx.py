#!/usr/bin/env python3
import argparse
import json
import sys
from pathlib import Path

import openpyxl

COL_JSON = 5
COL_X = 11
COL_Y = 12
COL_W = 13
COL_H = 14
COL_SCALE = 37


def to_float(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        value = value.strip()
        if value == "":
            return None
        try:
            return float(value)
        except ValueError:
            return None
    return None


def read_scale(json_path: Path):
    try:
        data = json.loads(json_path.read_text(encoding="utf-8"))
    except Exception as exc:
        return None, f"read json error: {exc}"

    width = to_float(data.get("width"))
    height = to_float(data.get("height"))
    if width is None or height is None:
        return None, "missing width/height"

    scale = max(width, height)
    if scale == 0:
        return None, "scale is 0"

    return scale, None


def resolve_json_path(root: Path, value):
    rel = Path(str(value))
    if rel.is_absolute():
        return rel
    return root / rel


def process_sheet(ws, root: Path):
    for row in range(2, ws.max_row + 1):
        cell_json = ws.cell(row=row, column=COL_JSON).value
        if cell_json is None or str(cell_json).strip() == "":
            continue

        json_path = resolve_json_path(root, cell_json)
        if not json_path.exists():
            print(f"row {row}: json not found: {json_path}", file=sys.stderr)
            continue

        scale, err = read_scale(json_path)
        if err:
            print(f"row {row}: {json_path} ({err})", file=sys.stderr)
            continue

        x = to_float(ws.cell(row=row, column=COL_X).value)
        y = to_float(ws.cell(row=row, column=COL_Y).value)
        w = to_float(ws.cell(row=row, column=COL_W).value)
        h = to_float(ws.cell(row=row, column=COL_H).value)
        if None in (x, y, w, h):
            print(f"row {row}: invalid x/y/w/h", file=sys.stderr)
            continue

        x = (x + w / 2.0) / scale
        y = (y + h / 2.0) / scale
        w = w / scale
        h = h / scale

        ws.cell(row=row, column=COL_X).value = x
        ws.cell(row=row, column=COL_Y).value = y
        ws.cell(row=row, column=COL_W).value = w
        ws.cell(row=row, column=COL_H).value = h
        ws.cell(row=row, column=COL_SCALE).value = scale


def main():
    parser = argparse.ArgumentParser(
        description="Normalize K/L/M/N using scale from JSON width/height."
    )
    parser.add_argument("xlsx", help="Path to input xlsx file.")
    parser.add_argument(
        "--root",
        help="Root directory for relative JSON paths (defaults to script directory).",
    )
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.is_file():
        print(f"Not a file: {xlsx_path}", file=sys.stderr)
        return 1

    root = Path(args.root) if args.root else Path(__file__).resolve().parent

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    process_sheet(ws, root)

    out_path = xlsx_path.with_name(xlsx_path.stem + "-n" + xlsx_path.suffix)
    wb.save(out_path)
    print(out_path)
    return 0


if __name__ == "__main__":
    sys.exit(main())
