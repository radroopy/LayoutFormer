#!/usr/bin/env python3
import argparse
import json
import math
import sys
from pathlib import Path

from PIL import Image, ImageDraw

try:
    import openpyxl
except ImportError:
    openpyxl = None

DEFAULT_LINE_WIDTH = 2
DEFAULT_BOX_WIDTH = 2

COL_JSON = 5
COL_X = 11
COL_Y = 12
COL_W = 13
COL_H = 14
COL_SCALE = 37
NORM_RANGE = 10.0

COLOR_ELEM_NORM = (0, 0, 0, 255)
COLOR_ELEM_ORIG = (80, 80, 80, 160)


def load_json(path: Path):
    return json.loads(path.read_text(encoding="utf-8"))


def pick_png(json_path: Path):
    stem = json_path.stem
    if stem.endswith("-n"):
        stem = stem[:-2]
    for ext in (".png", ".jpg", ".jpeg"):
        candidate = json_path.with_name(stem + ext)
        if candidate.exists():
            return candidate
    return None


def scale_point(pt, width, height, scale, img_w, img_h, normalized):
    if width == 0 or height == 0:
        return None
    x = pt[0]
    y = pt[1]
    if normalized:
        x *= scale / NORM_RANGE
        y *= scale / NORM_RANGE
    if img_w <= 1:
        x_px = 0.0
    else:
        x_px = x / width * (img_w - 1)
    if img_h <= 1:
        y_px = 0.0
    else:
        y_px = (1.0 - y / height) * (img_h - 1)
    return (x_px, y_px)


def point_to_pixel(x, y, width, height, img_w, img_h):
    if width == 0 or height == 0:
        return None
    if img_w <= 1:
        x_px = 0.0
    else:
        x_px = x / width * (img_w - 1)
    if img_h <= 1:
        y_px = 0.0
    else:
        y_px = (1.0 - y / height) * (img_h - 1)
    return (x_px, y_px)


def box_to_pixel(cx, cy, w, h, width, height, img_w, img_h):
    x0 = cx - w / 2.0
    y0 = cy - h / 2.0
    x1 = cx + w / 2.0
    y1 = cy + h / 2.0
    p0 = point_to_pixel(x0, y0, width, height, img_w, img_h)
    p1 = point_to_pixel(x1, y1, width, height, img_w, img_h)
    if p0 is None or p1 is None:
        return None
    return (p0[0], p0[1], p1[0], p1[1])


def choose_color(item):
    if item.get("isSewing"):
        return (0, 0, 255, 255)
    if item.get("isContour"):
        return (0, 0, 0, 255)
    if item.get("isGuides"):
        return (64, 64, 64, 255)
    return (64, 64, 64, 255)


def draw_path(draw, points, color, line_width, close=False):
    if len(points) < 2:
        return
    if close and points[0] != points[-1]:
        points = points + [points[0]]
    draw.line(points, fill=color, width=line_width)


def draw_boxes(draw, boxes, color, line_width):
    if not boxes:
        return
    lw = max(1, line_width)
    for box in boxes:
        if box is None:
            continue
        x0, y0, x1, y1 = box
        left = min(x0, x1)
        right = max(x0, x1)
        top = min(y0, y1)
        bottom = max(y0, y1)
        draw.line([(left, top), (right, top)], fill=color, width=lw)
        draw.line([(right, top), (right, bottom)], fill=color, width=lw)
        draw.line([(right, bottom), (left, bottom)], fill=color, width=lw)
        draw.line([(left, bottom), (left, top)], fill=color, width=lw)


def draw_item(draw, item, width, height, scale, img_w, img_h, line_width, normalized):
    color = choose_color(item)
    if isinstance(item.get("segments"), list) and item.get("segments"):
        for seg in item.get("segments"):
            if not isinstance(seg, list):
                continue
            pts = [
                scale_point(p, width, height, scale, img_w, img_h, normalized)
                for p in seg
                if isinstance(p, list) and len(p) == 2
            ]
            pts = [p for p in pts if p is not None]
            draw_path(draw, pts, color, line_width, close=False)
        return

    if isinstance(item.get("points"), list) and item.get("points"):
        pts = [
            scale_point(p, width, height, scale, img_w, img_h, normalized)
            for p in item.get("points")
            if isinstance(p, list) and len(p) == 2
        ]
        pts = [p for p in pts if p is not None]
        close = bool(item.get("isContour")) and len(pts) >= 3
        draw_path(draw, pts, color, line_width, close=close)


def to_float(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        value = value.strip()
        if value == "":
            return None
        try:
            return float(value)
        except ValueError:
            return None
    return None


def resolve_json_from_cell(root: Path, value):
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    path = Path(s)
    if path.suffix.lower() in (".png", ".jpg", ".jpeg"):
        if not path.is_absolute():
            path = root / path
        return path.with_suffix(".json")
    if path.is_absolute():
        return path
    return root / path


def read_scale_from_json(json_path: Path, cache):
    key = json_path.resolve()
    if key in cache:
        return cache[key]
    try:
        data = load_json(json_path)
    except Exception:
        cache[key] = None
        return None
    width = to_float(data.get("width"))
    height = to_float(data.get("height"))
    if width is None or height is None:
        cache[key] = None
        return None
    scale = max(width, height)
    cache[key] = scale
    return scale


def load_elements_map(xlsx_path: Path, root: Path, normalized: bool):
    if openpyxl is None:
        raise RuntimeError("openpyxl is required to read xlsx files")
    if not xlsx_path or not xlsx_path.exists():
        raise FileNotFoundError(f"xlsx not found: {xlsx_path}")

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    scale_cache = {}
    elements = {}

    for row in range(2, ws.max_row + 1):
        cell_ref = ws.cell(row=row, column=COL_JSON).value
        json_path = resolve_json_from_cell(root, cell_ref)
        if json_path is None:
            continue
        if not json_path.exists():
            print(f"row {row}: json not found: {json_path}", file=sys.stderr)
            continue

        x = to_float(ws.cell(row=row, column=COL_X).value)
        y = to_float(ws.cell(row=row, column=COL_Y).value)
        w = to_float(ws.cell(row=row, column=COL_W).value)
        h = to_float(ws.cell(row=row, column=COL_H).value)
        if None in (x, y, w, h):
            print(f"row {row}: invalid x/y/w/h", file=sys.stderr)
            continue

        if normalized:
            scale = to_float(ws.cell(row=row, column=COL_SCALE).value)
            if scale is None:
                scale = read_scale_from_json(json_path, scale_cache)
            if scale in (None, 0):
                print(f"row {row}: invalid scale", file=sys.stderr)
                continue
            x = x * scale / NORM_RANGE
            y = y * scale / NORM_RANGE
            w = w * scale / NORM_RANGE
            h = h * scale / NORM_RANGE
        else:
            x = x + w / 2.0
            y = y + h / 2.0

        key = json_path.resolve()
        elements.setdefault(key, []).append((x, y, w, h))

    return elements


def auto_pick_xlsx(base_dir: Path):
    files = [f for f in base_dir.glob("*.xlsx") if not f.name.startswith("~$")]
    norm = [f for f in files if f.stem.endswith("-n")]
    orig = [f for f in files if not f.stem.endswith("-n")]
    norm_path = norm[0] if len(norm) == 1 else None
    orig_path = orig[0] if len(orig) == 1 else None
    return orig_path, norm_path


def render_one(
    json_path: Path,
    png_path: Path,
    out_path: Path,
    line_width: int,
    normalized: bool,
    overlay: bool,
    max_side: int,
    elements_orig,
    elements_norm,
    box_width: int,
):
    data = load_json(json_path)
    width = data.get("width")
    height = data.get("height")

    if not isinstance(width, (int, float)) or not isinstance(height, (int, float)):
        print(f"skip {json_path}: invalid width/height", file=sys.stderr)
        return False

    scale = max(width, height)
    if scale == 0:
        print(f"skip {json_path}: scale is 0", file=sys.stderr)
        return False

    if overlay:
        image = Image.open(png_path).convert("RGBA")
        img_w, img_h = image.size
    else:
        factor = 1.0
        if max_side and max_side > 0:
            factor = max_side / max(width, height)
        img_w = max(1, int(math.ceil(width * factor)))
        img_h = max(1, int(math.ceil(height * factor)))
        image = Image.new("RGBA", (img_w, img_h), (255, 255, 255, 255))
    draw = ImageDraw.Draw(image)

    items = data.get("items")
    if isinstance(items, list):
        for item in items:
            if isinstance(item, dict):
                draw_item(
                    draw,
                    item,
                    width,
                    height,
                    scale,
                    img_w,
                    img_h,
                    line_width,
                    normalized,
                )

    json_key = json_path
    if json_key.stem.endswith("-n"):
        json_key = json_key.with_name(json_key.stem[:-2] + json_key.suffix)
    json_key = json_key.resolve()

    if elements_orig is not None:
        pts = elements_orig.get(json_key, [])
        boxes = [box_to_pixel(x, y, w, h, width, height, img_w, img_h) for x, y, w, h in pts]
        draw_boxes(draw, boxes, COLOR_ELEM_ORIG, box_width)

    if elements_norm is not None:
        pts = elements_norm.get(json_key, [])
        boxes = [box_to_pixel(x, y, w, h, width, height, img_w, img_h) for x, y, w, h in pts]
        draw_boxes(draw, boxes, COLOR_ELEM_NORM, box_width)

    image.save(out_path)
    print(out_path)
    return True


def iter_json_files(path: Path):
    if path.is_dir():
        yield from path.rglob("*-n.json")
    else:
        yield path


def main():
    parser = argparse.ArgumentParser(
        description="Rebuild contours from -n.json (bottom-left origin) or overlay on PNG."
    )
    parser.add_argument(
        "path",
        help="Path to a -n.json file or a directory (recursively scans *-n.json).",
    )
    parser.add_argument("--png", help="PNG path (only for single file mode).")
    parser.add_argument("--out", help="Output image path (only for single file mode).")
    parser.add_argument("--line-width", type=int, default=DEFAULT_LINE_WIDTH, help="Line width.")
    parser.add_argument(
        "--overlay",
        action="store_true",
        help="Overlay contours on the original PNG instead of a blank canvas.",
    )
    parser.add_argument(
        "--max-side",
        type=int,
        default=0,
        help="Max side length for rebuilt images (0 keeps original size).",
    )
    parser.add_argument(
        "--original",
        action="store_true",
        help="Treat points as original scale (skip normalization undo).",
    )
    parser.add_argument("--xlsx-orig", help="Original elements xlsx (non-normalized).")
    parser.add_argument("--xlsx-norm", help="Normalized elements xlsx (-n).")
    parser.add_argument(
        "--xlsx-root",
        help="Root dir for relative paths in column E (defaults to pattern/pattern).",
    )
    parser.add_argument(
        "--box-width",
        type=int,
        default=DEFAULT_BOX_WIDTH,
        help="Line width for element boxes.",
    )
    parser.add_argument(
        "--point-radius",
        dest="box_width",
        type=int,
        help="Alias for --box-width.",
    )
    args = parser.parse_args()

    path = Path(args.path)
    if not path.exists():
        print(f"Not found: {path}", file=sys.stderr)
        return 1

    normalized = not args.original

    base_dir = Path(__file__).resolve().parent
    default_root = base_dir / "pattern" / "pattern"
    if not default_root.exists():
        default_root = base_dir
    xlsx_root = Path(args.xlsx_root) if args.xlsx_root else default_root

    xlsx_orig = Path(args.xlsx_orig) if args.xlsx_orig else None
    xlsx_norm = Path(args.xlsx_norm) if args.xlsx_norm else None

    if xlsx_orig is None and xlsx_norm is None:
        xlsx_orig, xlsx_norm = auto_pick_xlsx(base_dir)

    elements_orig = None
    elements_norm = None
    if xlsx_orig or xlsx_norm:
        if openpyxl is None:
            print("openpyxl is required to read xlsx files", file=sys.stderr)
            return 1
        try:
            if xlsx_orig:
                elements_orig = load_elements_map(xlsx_orig, xlsx_root, normalized=False)
                print(f"loaded elements from {xlsx_orig}")
            if xlsx_norm:
                elements_norm = load_elements_map(xlsx_norm, xlsx_root, normalized=True)
                print(f"loaded elements from {xlsx_norm}")
        except Exception as exc:
            print(f"failed to load xlsx: {exc}", file=sys.stderr)
            return 1

    if path.is_file():
        if path.suffix.lower() != ".json" or not path.stem.endswith("-n"):
            print(f"Only -n.json files are supported: {path}", file=sys.stderr)
            return 1
        json_path = path
        png_path = None
        if args.overlay:
            png_path = Path(args.png) if args.png else pick_png(json_path)
            if not png_path or not png_path.exists():
                print(f"PNG not found for {json_path}", file=sys.stderr)
                return 1
        out_path = Path(args.out) if args.out else json_path.with_name(json_path.stem + "-vis.png")
        ok = render_one(
            json_path,
            png_path,
            out_path,
            args.line_width,
            normalized,
            args.overlay,
            args.max_side,
            elements_orig,
            elements_norm,
            args.box_width,
        )
        return 0 if ok else 1

    for json_path in iter_json_files(path):
        png_path = None
        if args.overlay:
            png_path = pick_png(json_path)
            if not png_path or not png_path.exists():
                print(f"PNG not found for {json_path}", file=sys.stderr)
                continue
        out_path = json_path.with_name(json_path.stem + "-vis.png")
        render_one(
            json_path,
            png_path,
            out_path,
            args.line_width,
            normalized,
            args.overlay,
            args.max_side,
            elements_orig,
            elements_norm,
            args.box_width,
        )

    return 0


if __name__ == "__main__":
    sys.exit(main())
