#!/usr/bin/env python3
import argparse
import json
import math
import sys
from pathlib import Path

import openpyxl

COL_JSON = 5
COL_X = 11
COL_Y = 12
COL_W = 13
COL_H = 14
COL_SCALE = 37

DEFAULT_EPS = 1e-6
NORM_RANGE = 10.0


class ErrorTracker:
    def __init__(self, max_errors):
        self.max_errors = max_errors
        self.count = 0
        self.stopped = False

    def add(self, msg):
        self.count += 1
        print(msg, file=sys.stderr)
        if self.count >= self.max_errors:
            print("error limit reached", file=sys.stderr)
            self.stopped = True


def to_float(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        value = value.strip()
        if value == "":
            return None
        try:
            return float(value)
        except ValueError:
            return None
    return None


def float_close(a, b, eps):
    return math.isclose(a, b, rel_tol=eps, abs_tol=eps)


def read_json(path: Path):
    return json.loads(path.read_text(encoding="utf-8"))


def strip_points(obj):
    if isinstance(obj, dict):
        return {k: strip_points(v) for k, v in obj.items() if k not in ("points", "segments")}
    if isinstance(obj, list):
        return [strip_points(v) for v in obj]
    return obj


def expected_point(pt, scale):
    return [pt[0] / scale * NORM_RANGE, pt[1] / scale * NORM_RANGE]


def validate_point_list(orig, norm, scale, eps, tracker, ctx):
    if not isinstance(orig, list) or not isinstance(norm, list):
        tracker.add(f"{ctx}: points not list")
        return
    if len(orig) != len(norm):
        tracker.add(f"{ctx}: points len mismatch {len(orig)} != {len(norm)}")
        return
    for i, (p_orig, p_norm) in enumerate(zip(orig, norm)):
        if tracker.stopped:
            return
        if not (isinstance(p_orig, list) and isinstance(p_norm, list) and len(p_orig) == 2 and len(p_norm) == 2):
            tracker.add(f"{ctx}: invalid point at {i}")
            return
        exp = expected_point(p_orig, scale)
        if not (float_close(p_norm[0], exp[0], eps) and float_close(p_norm[1], exp[1], eps)):
            tracker.add(
                f"{ctx}: point {i} mismatch expected ({exp[0]}, {exp[1]}) got ({p_norm[0]}, {p_norm[1]})"
            )
            return


def validate_segments(orig, norm, scale, eps, tracker, ctx):
    if not isinstance(orig, list) or not isinstance(norm, list):
        tracker.add(f"{ctx}: segments not list")
        return
    if len(orig) != len(norm):
        tracker.add(f"{ctx}: segments len mismatch {len(orig)} != {len(norm)}")
        return
    for s_idx, (seg_orig, seg_norm) in enumerate(zip(orig, norm)):
        if tracker.stopped:
            return
        validate_point_list(seg_orig, seg_norm, scale, eps, tracker, f"{ctx} seg {s_idx}")
        if tracker.stopped:
            return


def validate_json_files(root: Path, recursive: bool, eps: float, tracker: ErrorTracker):
    pattern = "*-n.json"
    files = root.rglob(pattern) if recursive else root.glob(pattern)
    checked = 0
    for norm_path in files:
        if tracker.stopped:
            break
        stem = norm_path.stem
        if not stem.endswith("-n"):
            continue
        orig_path = norm_path.with_name(stem[:-2] + norm_path.suffix)
        if not orig_path.exists():
            tracker.add(f"json: missing original for {norm_path}")
            continue
        try:
            orig = read_json(orig_path)
            norm = read_json(norm_path)
        except Exception as exc:
            tracker.add(f"json: read error {norm_path} ({exc})")
            continue

        scale = max(to_float(orig.get("width")) or 0, to_float(orig.get("height")) or 0)
        if scale == 0:
            tracker.add(f"json: invalid scale in {orig_path}")
            continue

        if strip_points(orig) != strip_points(norm):
            tracker.add(f"json: non-point fields differ {norm_path}")
            continue

        orig_items = orig.get("items")
        norm_items = norm.get("items")
        if not isinstance(orig_items, list) or not isinstance(norm_items, list):
            tracker.add(f"json: items not list {norm_path}")
            continue
        if len(orig_items) != len(norm_items):
            tracker.add(f"json: items len mismatch {norm_path}")
            continue

        for i, (orig_item, norm_item) in enumerate(zip(orig_items, norm_items)):
            if tracker.stopped:
                break
            if "points" in orig_item or "points" in norm_item:
                validate_point_list(
                    orig_item.get("points"),
                    norm_item.get("points"),
                    scale,
                    eps,
                    tracker,
                    f"{norm_path} item {i} points",
                )
            if tracker.stopped:
                break
            if "segments" in orig_item or "segments" in norm_item:
                validate_segments(
                    orig_item.get("segments"),
                    norm_item.get("segments"),
                    scale,
                    eps,
                    tracker,
                    f"{norm_path} item {i} segments",
                )

        checked += 1

    return checked


def resolve_json_path(root: Path, value):
    rel = Path(str(value))
    if rel.is_absolute():
        return rel
    return root / rel


def read_scale(json_path: Path):
    try:
        data = json.loads(json_path.read_text(encoding="utf-8"))
    except Exception as exc:
        return None, f"read json error: {exc}"

    width = to_float(data.get("width"))
    height = to_float(data.get("height"))
    if width is None or height is None:
        return None, "missing width/height"

    scale = max(width, height)
    if scale == 0:
        return None, "scale is 0"

    return scale, None


def validate_xlsx(xlsx_path: Path, norm_path: Path, root: Path, eps: float, tracker: ErrorTracker):
    if not norm_path.exists():
        tracker.add(f"xlsx: normalized file not found {norm_path}")
        return 0

    wb_orig = openpyxl.load_workbook(xlsx_path)
    wb_norm = openpyxl.load_workbook(norm_path)
    ws_orig = wb_orig.active
    ws_norm = wb_norm.active

    max_row = max(ws_orig.max_row, ws_norm.max_row)
    max_col = max(ws_orig.max_column, ws_norm.max_column)

    rows_checked = 0
    for row in range(2, max_row + 1):
        if tracker.stopped:
            break
        cell_json = ws_orig.cell(row=row, column=COL_JSON).value
        if cell_json is None or str(cell_json).strip() == "":
            for col in range(1, max_col + 1):
                if tracker.stopped:
                    break
                v_orig = ws_orig.cell(row=row, column=col).value
                v_norm = ws_norm.cell(row=row, column=col).value
                if v_orig != v_norm:
                    tracker.add(f"xlsx: row {row} col {col} changed but no json path")
                    break
            rows_checked += 1
            continue

        json_path = resolve_json_path(root, cell_json)
        if not json_path.exists():
            tracker.add(f"xlsx: row {row} json not found {json_path}")
            rows_checked += 1
            continue

        scale, err = read_scale(json_path)
        if err:
            tracker.add(f"xlsx: row {row} {json_path} ({err})")
            rows_checked += 1
            continue

        x = to_float(ws_orig.cell(row=row, column=COL_X).value)
        y = to_float(ws_orig.cell(row=row, column=COL_Y).value)
        w = to_float(ws_orig.cell(row=row, column=COL_W).value)
        h = to_float(ws_orig.cell(row=row, column=COL_H).value)
        if None in (x, y, w, h):
            tracker.add(f"xlsx: row {row} invalid x/y/w/h")
            rows_checked += 1
            continue

        exp_x = (x + w / 2.0) / scale * NORM_RANGE
        exp_y = (y + h / 2.0) / scale * NORM_RANGE
        exp_w = w / scale * NORM_RANGE
        exp_h = h / scale * NORM_RANGE

        got_x = to_float(ws_norm.cell(row=row, column=COL_X).value)
        got_y = to_float(ws_norm.cell(row=row, column=COL_Y).value)
        got_w = to_float(ws_norm.cell(row=row, column=COL_W).value)
        got_h = to_float(ws_norm.cell(row=row, column=COL_H).value)
        got_scale = to_float(ws_norm.cell(row=row, column=COL_SCALE).value)

        if None in (got_x, got_y, got_w, got_h, got_scale):
            tracker.add(f"xlsx: row {row} missing normalized values")
            rows_checked += 1
            continue

        if not (
            float_close(got_x, exp_x, eps)
            and float_close(got_y, exp_y, eps)
            and float_close(got_w, exp_w, eps)
            and float_close(got_h, exp_h, eps)
            and float_close(got_scale, scale, eps)
        ):
            tracker.add(f"xlsx: row {row} normalized values mismatch")
            rows_checked += 1
            continue

        for col in range(1, max_col + 1):
            if tracker.stopped:
                break
            if col in (COL_X, COL_Y, COL_W, COL_H, COL_SCALE):
                continue
            v_orig = ws_orig.cell(row=row, column=col).value
            v_norm = ws_norm.cell(row=row, column=col).value
            if v_orig != v_norm:
                tracker.add(f"xlsx: row {row} col {col} changed unexpectedly")
                break

        rows_checked += 1

    return rows_checked


def main():
    parser = argparse.ArgumentParser(
        description="Validate normalized -n.json and -n.xlsx outputs."
    )
    parser.add_argument("--json-dir", help="Root directory to scan for *-n.json.")
    parser.add_argument("--no-recursive", action="store_true", help="Do not scan subdirectories.")
    parser.add_argument("--xlsx", help="Path to original xlsx file.")
    parser.add_argument("--normalized-xlsx", help="Path to normalized xlsx file.")
    parser.add_argument(
        "--root",
        help="Root directory for relative JSON paths (defaults to script directory).",
    )
    parser.add_argument("--eps", type=float, default=DEFAULT_EPS, help="Tolerance for floats.")
    parser.add_argument("--max-errors", type=int, default=50, help="Stop after this many errors.")
    args = parser.parse_args()

    if not args.json_dir and not args.xlsx:
        print("Provide --json-dir and/or --xlsx", file=sys.stderr)
        return 1

    tracker = ErrorTracker(args.max_errors)
    root = Path(args.root) if args.root else Path(__file__).resolve().parent

    if args.json_dir:
        json_root = Path(args.json_dir)
        if not json_root.is_dir():
            tracker.add(f"json: not a directory {json_root}")
        else:
            checked = validate_json_files(json_root, not args.no_recursive, args.eps, tracker)
            print(f"json files checked: {checked}")

    if args.xlsx:
        xlsx_path = Path(args.xlsx)
        if not xlsx_path.is_file():
            tracker.add(f"xlsx: not a file {xlsx_path}")
        else:
            norm_path = Path(args.normalized_xlsx) if args.normalized_xlsx else xlsx_path.with_name(
                xlsx_path.stem + "-n" + xlsx_path.suffix
            )
            rows_checked = validate_xlsx(xlsx_path, norm_path, root, args.eps, tracker)
            print(f"xlsx rows checked: {rows_checked}")

    if tracker.count:
        print(f"errors: {tracker.count}", file=sys.stderr)
        return 1

    print("validation passed")
    return 0


if __name__ == "__main__":
    sys.exit(main())
