#!/usr/bin/env python3
import argparse
import json
import sys
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path

try:
    import openpyxl
    from openpyxl.utils import column_index_from_string
except ImportError as exc:
    raise SystemExit("openpyxl is required to run this script") from exc

COL_ID = 1      # A
COL_JSON = 5    # E
COL_PDF = 6     # F
DEFAULT_TYPE_COL = 8  # H


def parse_column(value):
    if value is None:
        return None
    if isinstance(value, int):
        return value
    text = str(value).strip()
    if text == "":
        return None
    if text.isdigit():
        return int(text)
    try:
        return column_index_from_string(text.upper())
    except Exception:
        return None


def _make_rel_id(path: Path, project_root: Path, fallback_root: Path | None = None) -> str:
    try:
        return path.resolve().relative_to(project_root).as_posix()
    except Exception:
        if fallback_root is not None:
            try:
                return path.resolve().relative_to(fallback_root).as_posix()
            except Exception:
                pass
        return path.resolve().as_posix()


def resolve_path(root: Path, value) -> Path:
    rel = Path(str(value))
    if rel.is_absolute():
        return rel
    return root / rel


def extract_style_size_piece(json_rel: str):
    parts = Path(json_rel).parts
    if len(parts) < 3:
        return None, None, Path(json_rel).stem
    style = parts[-3]
    size = parts[-2]
    piece = Path(parts[-1]).stem
    return style, size, piece


def main():
    parser = argparse.ArgumentParser(
        description="Group pattern pieces by pdf<->json connectivity and report conflicts."
    )
    parser.add_argument("xlsx", help="Path to input xlsx file.")
    parser.add_argument(
        "--root",
        help="Root directory for relative JSON paths (defaults to project_root/pattern/pattern if exists).",
    )
    parser.add_argument(
        "--type-col",
        help="Type column (letter or 1-based index). Default: H. Used only for stats; not required.",
    )
    parser.add_argument(
        "--out",
        help="Output json path (default: <project>/data/shape_element_map.json).",
    )
    parser.add_argument(
        "--out-index",
        help="Output shape-size index json (default: <project>/data/shape_size_index.json).",
    )
    parser.add_argument(
        "--out-json-lookup",
        help="Output json->size map (default: <project>/data/json_size_lookup.json).",
    )
    parser.add_argument(
        "--out-pdf-map",
        help="Output pdf->shape map (default: <project>/data/pdf_shape_map.json).",
    )
    parser.add_argument(
        "--out-conflicts",
        help="Output conflict report path (default: <project>/data/shape_conflict_report.txt).",
    )
    args = parser.parse_args()

    project_root = Path(__file__).resolve().parent
    default_root = project_root / "pattern" / "pattern"
    if not default_root.is_dir():
        default_root = project_root
    root = Path(args.root) if args.root else default_root

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.is_file():
        print(f"Not a file: {xlsx_path}", file=sys.stderr)
        return 1

    out_path = Path(args.out) if args.out else project_root / "data" / "shape_element_map.json"
    out_path.parent.mkdir(parents=True, exist_ok=True)

    out_index = Path(args.out_index) if args.out_index else project_root / "data" / "shape_size_index.json"
    out_index.parent.mkdir(parents=True, exist_ok=True)

    out_json_lookup = Path(args.out_json_lookup) if args.out_json_lookup else project_root / "data" / "json_size_lookup.json"
    out_json_lookup.parent.mkdir(parents=True, exist_ok=True)

    out_pdf_map = Path(args.out_pdf_map) if args.out_pdf_map else project_root / "data" / "pdf_shape_map.json"
    out_pdf_map.parent.mkdir(parents=True, exist_ok=True)

    out_conflicts = Path(args.out_conflicts) if args.out_conflicts else project_root / "data" / "shape_conflict_report.txt"
    out_conflicts.parent.mkdir(parents=True, exist_ok=True)

    type_col = parse_column(args.type_col) if args.type_col else DEFAULT_TYPE_COL

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    pattern_map = {}
    pdf_size_map = defaultdict(dict)
    pdf_to_jsons = defaultdict(set)
    json_to_pdfs = defaultdict(set)
    size_conflicts = []
    skipped = 0

    for row in range(2, ws.max_row + 1):
        element_id = ws.cell(row=row, column=COL_ID).value
        json_cell = ws.cell(row=row, column=COL_JSON).value
        pdf_cell = ws.cell(row=row, column=COL_PDF).value
        type_cell = ws.cell(row=row, column=type_col).value if type_col else None

        if element_id is None or str(element_id).strip() == "":
            continue
        if json_cell is None or str(json_cell).strip() == "":
            skipped += 1
            continue
        if pdf_cell is None or str(pdf_cell).strip() == "":
            skipped += 1
            continue

        json_path = resolve_path(root, json_cell)
        json_rel = _make_rel_id(json_path, project_root, fallback_root=root)

        pdf_path = resolve_path(project_root, pdf_cell)
        pdf_rel = _make_rel_id(pdf_path, project_root, fallback_root=project_root)

        type_name = None
        if type_cell is not None:
            text = str(type_cell).strip()
            if text != "":
                type_name = text

        style, size, piece = extract_style_size_piece(json_rel)

        entry = pattern_map.setdefault(json_rel, {
            "json_path": str(json_path),
            "pdf_paths": set(),
            "style": style,
            "size": size,
            "piece": piece,
            "element_ids": [],
            "types": [],
        })
        entry["pdf_paths"].add(pdf_rel)
        entry["element_ids"].append(str(element_id).strip())
        if type_name is not None:
            entry["types"].append(type_name)

        if size:
            if size in pdf_size_map[pdf_rel] and pdf_size_map[pdf_rel][size] != json_rel:
                size_conflicts.append({
                    "pdf": pdf_rel,
                    "size": size,
                    "existing": pdf_size_map[pdf_rel][size],
                    "incoming": json_rel,
                })
            else:
                pdf_size_map[pdf_rel][size] = json_rel

        pdf_to_jsons[pdf_rel].add(json_rel)
        json_to_pdfs[json_rel].add(pdf_rel)

    # Build shapes using connectivity (pdf <-> json)
    visited_pdfs = set()
    visited_jsons = set()
    shapes = {}
    pdf_to_shape = {}
    shape_size_index = {}
    shape_conflicts = []
    shape_coverage_issues = []

    for start_pdf in sorted(pdf_to_jsons.keys()):
        if start_pdf in visited_pdfs:
            continue

        queue = [("pdf", start_pdf)]
        pdfs = set()
        jsons = set()

        while queue:
            kind, node = queue.pop()
            if kind == "pdf":
                if node in visited_pdfs:
                    continue
                visited_pdfs.add(node)
                pdfs.add(node)
                for json_rel in pdf_to_jsons.get(node, []):
                    if json_rel not in visited_jsons:
                        queue.append(("json", json_rel))
            else:
                if node in visited_jsons:
                    continue
                visited_jsons.add(node)
                jsons.add(node)
                for pdf in json_to_pdfs.get(node, []):
                    if pdf not in visited_pdfs:
                        queue.append(("pdf", pdf))

        if not pdfs:
            continue

        pdfs_sorted = sorted(pdfs)
        jsons_sorted = sorted(jsons)
        shape_id = pdfs_sorted[0]

        size_entries = defaultdict(list)
        size_sets = {}
        for pdf in pdfs_sorted:
            size_map_for_pdf = pdf_size_map.get(pdf, {})
            size_sets[pdf] = set(size_map_for_pdf.keys())
            for size, json_rel in size_map_for_pdf.items():
                size_entries[size].append({"pdf": pdf, "json": json_rel})

        all_sizes = set()
        for sizes in size_sets.values():
            all_sizes.update(sizes)
        if size_sets:
            if len(size_sets) > 1:
                common_sizes = set.intersection(*size_sets.values())
            else:
                common_sizes = set(next(iter(size_sets.values())))
        else:
            common_sizes = set()

        size_map = {}
        for size, entries in size_entries.items():
            entries_sorted = sorted(entries, key=lambda e: e["pdf"])
            size_map[size] = entries_sorted[0]["json"]
            unique_jsons = sorted({e["json"] for e in entries_sorted})
            if len(unique_jsons) > 1:
                shape_conflicts.append({
                    "shape_id": shape_id,
                    "size": size,
                    "jsons": unique_jsons,
                    "entries": entries_sorted,
                })

        coverage = []
        for pdf in pdfs_sorted:
            sizes = size_sets.get(pdf, set())
            missing_sizes = sorted(all_sizes - sizes)
            extra_sizes = sorted(sizes - common_sizes)
            if missing_sizes or extra_sizes:
                coverage.append({
                    "pdf": pdf,
                    "missing_sizes": missing_sizes,
                    "extra_sizes": extra_sizes,
                })
        if coverage:
            shape_coverage_issues.append({
                "shape_id": shape_id,
                "all_sizes": sorted(all_sizes),
                "coverage": coverage,
            })

        shapes[shape_id] = {
            "pdfs": pdfs_sorted,
            "sizes": size_map,
            "jsons": jsons_sorted,
        }
        shape_size_index[shape_id] = size_map
        for pdf in pdfs_sorted:
            pdf_to_shape[pdf] = shape_id

    # Build json_size_lookup (one-step lookup)
    json_size_lookup = {}
    for json_rel in pattern_map.keys():
        pdfs = json_to_pdfs.get(json_rel, [])
        shape_id = None
        for pdf in pdfs:
            shape_id = pdf_to_shape.get(pdf)
            if shape_id:
                break
        if not shape_id:
            continue
        size_map = shape_size_index.get(shape_id)
        if size_map is None:
            continue
        json_size_lookup[json_rel] = size_map

    # type counts for patterns
    patterns_out = {}
    for json_rel, info in pattern_map.items():
        type_counts = Counter(info["types"])
        pdfs_sorted = sorted(info["pdf_paths"])
        patterns_out[json_rel] = {
            "pdf_paths": pdfs_sorted,
            "pdf_path": pdfs_sorted[0] if pdfs_sorted else None,
            "style": info.get("style"),
            "size": info.get("size"),
            "piece": info.get("piece"),
            "element_count": len(info["element_ids"]),
            "type_counts": dict(type_counts),
            "element_ids": info.get("element_ids"),
        }

    out_index.write_text(json.dumps(shape_size_index, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    out_json_lookup.write_text(json.dumps(json_size_lookup, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    out_pdf_map.write_text(json.dumps(pdf_to_shape, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    if shape_conflicts or shape_coverage_issues:
        lines = []
        if shape_conflicts:
            lines.append("[size->json conflicts]")
            for conflict in shape_conflicts:
                lines.append(f"- shape_id: {conflict['shape_id']}")
                lines.append(f"  size: {conflict['size']}")
                lines.append(f"  jsons: {conflict['jsons']}")
                lines.append("  entries:")
                for entry in conflict["entries"]:
                    lines.append(f"    {entry['pdf']} -> {entry['json']}")
        if shape_coverage_issues:
            if lines:
                lines.append("")
            lines.append("[shape coverage issues]")
            for issue in shape_coverage_issues:
                lines.append(f"- shape_id: {issue['shape_id']}")
                lines.append(f"  all_sizes: {issue['all_sizes']}")
                lines.append("  pdfs:")
                for item in issue["coverage"]:
                    lines.append(f"    {item['pdf']}")
                    if item["missing_sizes"]:
                        lines.append(f"      missing_sizes: {item['missing_sizes']}")
                    if item["extra_sizes"]:
                        lines.append(f"      extra_sizes: {item['extra_sizes']}")
        out_conflicts.write_text("\n".join(lines) + "\n", encoding="utf-8")
    else:
        out_conflicts.write_text("(no conflicts)\n", encoding="utf-8")

    result = {
        "created_at": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
        "project_root": str(project_root),
        "json_root": str(root),
        "xlsx": str(xlsx_path),
        "type_col": int(type_col) if type_col else None,
        "skipped_rows": skipped,
        "size_conflicts": size_conflicts,
        "shape_conflicts": shape_conflicts,
        "shape_coverage_issues": shape_coverage_issues,
        "shape_size_index_path": str(out_index),
        "json_size_lookup_path": str(out_json_lookup),
        "pdf_shape_map_path": str(out_pdf_map),
        "conflict_report_path": str(out_conflicts),
        "patterns": patterns_out,
        "shapes": shapes,
    }

    out_path.write_text(json.dumps(result, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    print(f"pdfs: {len(pdf_to_jsons)}")
    print(f"shapes: {len(shapes)}")
    print(f"size_conflicts: {len(size_conflicts)}")
    print(f"shape_conflicts: {len(shape_conflicts)}")
    print(f"shape_coverage_issues: {len(shape_coverage_issues)}")
    print(f"shape_size_index: {out_index}")
    print(f"json_size_lookup: {out_json_lookup}")
    print(f"pdf_shape_map: {out_pdf_map}")
    print(f"conflict_report: {out_conflicts}")
    print(f"output: {out_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
