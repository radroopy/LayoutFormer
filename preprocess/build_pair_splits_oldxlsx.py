#!/usr/bin/env python3
import argparse
import json
import random
import sys
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
try:
    import numpy as np
except ImportError:
    np = None

ROOT = Path(__file__).resolve().parents[1]
COL_SAMPLE = 2     # column B
COL_SHAPE_NAME = 4  # column D
COL_JSON = 5        # column E


def shape_id_from_row(sample_id, shape_name, json_path) -> str | None:
    if sample_id is None or str(sample_id).strip() == "":
        return None
    sample_text = str(sample_id).strip()
    if shape_name is None or str(shape_name).strip() == "":
        return None
    name = str(shape_name).strip()
    if json_path is None or str(json_path).strip() == "":
        return None
    parts = Path(str(json_path)).parts
    style = parts[-3] if len(parts) >= 3 else "UNKNOWN"
    return f"{sample_text}-{style}-{name}"


def to_str(value):
    if isinstance(value, bytes):
        return value.decode("utf-8")
    return str(value)


def load_valid_layouts(npz_path: Path) -> dict | None:
    if not npz_path.is_file():
        return None
    if np is None:
        raise SystemExit("numpy is required to read elements_embed.npz")
    data = np.load(npz_path, allow_pickle=True)
    if "shape_ids" not in data or "pattern_json_paths" not in data:
        print(f"[WARN] elements_embed missing keys: {npz_path}", file=sys.stderr)
        return None
    shape_ids = [to_str(v) for v in data["shape_ids"].tolist()]
    json_paths = [to_str(v) for v in data["pattern_json_paths"].tolist()]
    valid = {}
    for shape_id, json_path in zip(shape_ids, json_paths):
        if not shape_id or not json_path:
            continue
        valid.setdefault(shape_id, set()).add(json_path)
    return valid


def load_shape_pdf_map(path: Path) -> dict | None:
    if not path.is_file():
        return None
    data = json.loads(path.read_text(encoding="utf-8"))
    shapes = data.get("shapes", {})
    if not isinstance(shapes, dict):
        return None
    mapping = {}
    for shape_id, info in shapes.items():
        json_elements = info.get("json_elements", {})
        if not isinstance(json_elements, dict):
            continue
        pdf_map = {}
        for json_path, entry in json_elements.items():
            if not isinstance(entry, dict):
                continue
            pdfs = entry.get("pdfs", [])
            if isinstance(pdfs, list):
                pdf_map[json_path] = list(pdfs)
        if pdf_map:
            mapping[shape_id] = pdf_map
    return mapping


def collect_json_wh(scale_payload: dict) -> dict:
    json_wh = {}
    results = scale_payload.get("results", {})
    if not isinstance(results, dict):
        return json_wh

    def add_entry(entry):
        json_path = entry.get("json")
        if not json_path:
            return
        w = float(entry.get("width", 0.0))
        h = float(entry.get("height", 0.0))
        if json_path not in json_wh:
            json_wh[json_path] = (w, h)

    old_format = bool(results) and all(
        isinstance(v, dict) and "sizes" in v for v in results.values()
    )
    if old_format:
        for _, info in results.items():
            for _, entry in info.get("sizes", {}).items():
                add_entry(entry)
        return json_wh

    for _, shape_results in results.items():
        if not isinstance(shape_results, dict):
            continue
        for _, info in shape_results.items():
            if not isinstance(info, dict):
                continue
            for _, entry in info.get("sizes", {}).items():
                add_entry(entry)
    return json_wh


def collect_test_shapes(old_xlsx: Path) -> set[str]:
    wb = openpyxl.load_workbook(old_xlsx)
    ws = wb.active
    shapes = set()
    for row in range(2, ws.max_row + 1):
        sample_id = ws.cell(row=row, column=COL_SAMPLE).value
        shape_name = ws.cell(row=row, column=COL_SHAPE_NAME).value
        json_path = ws.cell(row=row, column=COL_JSON).value
        sid = shape_id_from_row(sample_id, shape_name, json_path)
        if sid:
            shapes.add(sid)
    return shapes


def main():
    parser = argparse.ArgumentParser(
        description="Build pair splits using old xlsx as TEST shapes; remaining shapes -> train/val."
    )
    parser.add_argument(
        "--shape-index",
        default=str(ROOT / "data" / "shape_size_index.json"),
        help="Path to shape_size_index.json",
    )
    parser.add_argument(
        "--size-scales",
        default=str(ROOT / "data" / "size_scale_factors.json"),
        help="Path to size_scale_factors.json (for target width/height/scale)",
    )
    parser.add_argument(
        "--old-xlsx",
        required=True,
        help="Old xlsx used to define TEST shapes.",
    )
    parser.add_argument(
        "--out",
        default=str(ROOT / "train" / "pair_splits.json"),
        help="Output json path for splits",
    )
    parser.add_argument(
        "--elements-embed",
        default=str(ROOT / "data" / "element_emb" / "elements_embed.npz"),
        help="Path to elements_embed.npz (used to filter jsons with no valid elements).",
    )
    parser.add_argument(
        "--shape-elements",
        default=str(ROOT / "data" / "shape_element_map.json"),
        help="Path to shape_element_map.json (used to filter pairs with pdf mismatch).",
    )
    parser.add_argument("--seed", type=int, default=42, help="Random seed")
    parser.add_argument("--train-ratio", type=float, default=0.95, help="Train ratio")
    parser.add_argument("--val-ratio", type=float, default=0.05, help="Val ratio")
    parser.add_argument(
        "--exclude-self",
        action="store_true",
        help="Exclude same-size pairs (src_size==tgt_size). Default is to include them (e.g., L->L).",
    )
    args = parser.parse_args()

    shape_index_path = Path(args.shape_index)
    if not shape_index_path.is_file():
        raise SystemExit(f"Not found: {shape_index_path}")

    size_scale_path = Path(args.size_scales)
    json_wh = {}
    if size_scale_path.is_file():
        scale_payload = json.loads(size_scale_path.read_text(encoding="utf-8"))
        json_wh = collect_json_wh(scale_payload)

    old_xlsx_path = Path(args.old_xlsx)
    if not old_xlsx_path.is_file():
        raise SystemExit(f"old_xlsx not found: {old_xlsx_path}")
    test_shapes = collect_test_shapes(old_xlsx_path)

    shapes = json.loads(shape_index_path.read_text(encoding="utf-8"))
    rng = random.Random(args.seed)

    elements_path = Path(args.elements_embed)
    valid_layouts = load_valid_layouts(elements_path)
    if valid_layouts is None:
        if elements_path.is_file():
            print(f"[WARN] skip filtering: invalid elements_embed {elements_path}", file=sys.stderr)
        else:
            print(f"[WARN] skip filtering: elements_embed not found {elements_path}", file=sys.stderr)
    dropped_jsons = 0
    dropped_shapes = 0
    dropped_shape_reasons = {
        "missing_elements": 0,
        "missing_shape_pdfs": 0,
        "missing_json_pdfs": 0,
        "shape_element_count_mismatch": 0,
    }

    shape_elements_path = Path(args.shape_elements)
    shape_pdf_map = load_shape_pdf_map(shape_elements_path)
    if shape_pdf_map is None:
        if shape_elements_path.is_file():
            print(f"[WARN] skip pdf filter: invalid shape_element_map {shape_elements_path}", file=sys.stderr)
        else:
            print(f"[WARN] skip pdf filter: shape_element_map not found {shape_elements_path}", file=sys.stderr)
    dropped_pairs = 0

    splits = {"train": [], "val": [], "test": []}
    shape_stats = {}

    for shape_id, size_map in shapes.items():
        if valid_layouts is not None:
            valid_jsons = valid_layouts.get(shape_id, set())
            if not valid_jsons:
                for size, json_path in size_map.items():
                    print(
                        f"[drop] shape={shape_id} size={size} json={json_path} reason=missing_elements",
                        file=sys.stderr,
                    )
                dropped_jsons += len(size_map)
                dropped_shapes += 1
                dropped_shape_reasons["missing_elements"] += 1
                continue
            filtered_map = {}
            for size, json_path in size_map.items():
                if json_path in valid_jsons:
                    filtered_map[size] = json_path
                else:
                    dropped_jsons += 1
                    print(
                        f"[drop] shape={shape_id} size={size} json={json_path} reason=missing_elements",
                        file=sys.stderr,
                    )
            size_map = filtered_map
            if not size_map:
                dropped_shapes += 1
                dropped_shape_reasons["missing_elements"] += 1
                continue

        if shape_pdf_map is not None:
            pdfs_by_json = shape_pdf_map.get(shape_id)
            if not pdfs_by_json:
                for size, json_path in size_map.items():
                    print(
                        f"[drop] shape={shape_id} size={size} json={json_path} reason=missing_shape_pdfs",
                        file=sys.stderr,
                    )
                dropped_jsons += len(size_map)
                dropped_shapes += 1
                dropped_shape_reasons["missing_shape_pdfs"] += 1
                continue

            size_counts = {}
            missing = []
            for size, json_path in size_map.items():
                pdfs = pdfs_by_json.get(json_path)
                if not pdfs:
                    missing.append((size, json_path))
                    continue
                size_counts[size] = len(pdfs)

            if missing:
                dropped_jsons += len(size_map)
                dropped_shapes += 1
                dropped_shape_reasons["missing_json_pdfs"] += 1
                missing_desc = ", ".join(f"{size}:{json_path}" for size, json_path in missing)
                print(
                    f"[drop] shape={shape_id} reason=missing_json_pdfs missing={missing_desc}",
                    file=sys.stderr,
                )
                continue

            if size_counts:
                unique_counts = sorted(set(size_counts.values()))
                if len(unique_counts) > 1:
                    dropped_jsons += len(size_map)
                    dropped_shapes += 1
                    dropped_shape_reasons["shape_element_count_mismatch"] += 1
                    counts_desc = ", ".join(f"{k}:{v}" for k, v in sorted(size_counts.items()))
                    print(
                        f"[drop] shape={shape_id} reason=shape_element_count_mismatch counts={counts_desc}",
                        file=sys.stderr,
                    )
                    continue

        sizes = sorted(size_map.keys())
        pairs = []
        for src_size in sizes:
            for tgt_size in sizes:
                if args.exclude_self and src_size == tgt_size:
                    continue
                src_json = size_map[src_size]
                tgt_json = size_map[tgt_size]
                if shape_pdf_map is not None:
                    pdfs_by_json = shape_pdf_map.get(shape_id)
                    if not pdfs_by_json:
                        dropped_pairs += 1
                        print(
                            f"[drop] shape={shape_id} src={src_json} tgt={tgt_json} reason=missing_shape_pdfs",
                            file=sys.stderr,
                        )
                        continue
                    src_pdfs = pdfs_by_json.get(src_json)
                    tgt_pdfs = pdfs_by_json.get(tgt_json)
                    if not src_pdfs or not tgt_pdfs:
                        dropped_pairs += 1
                        print(
                            f"[drop] shape={shape_id} src={src_json} tgt={tgt_json} reason=missing_json_pdfs",
                            file=sys.stderr,
                        )
                        continue
                    if src_pdfs != tgt_pdfs:
                        dropped_pairs += 1
                        print(
                            f"[drop] shape={shape_id} src={src_json} tgt={tgt_json} "
                            f"reason=pdf_mismatch src_n={len(src_pdfs)} tgt_n={len(tgt_pdfs)}",
                            file=sys.stderr,
                        )
                        continue
                w_h = json_wh.get(tgt_json)
                tgt_w = None
                tgt_h = None
                tgt_scale = None
                if w_h is not None:
                    tgt_w, tgt_h = w_h
                    tgt_scale = max(tgt_w, tgt_h)
                pairs.append(
                    {
                        "shape_id": shape_id,
                        "src_size": src_size,
                        "tgt_size": tgt_size,
                        "src_json": src_json,
                        "tgt_json": tgt_json,
                        "tgt_width": tgt_w,
                        "tgt_height": tgt_h,
                        "tgt_scale": tgt_scale,
                    }
                )

        rng.shuffle(pairs)
        n = len(pairs)

        if shape_id in test_shapes:
            splits["test"].extend(pairs)
            n_train = 0
            n_val = 0
            n_test = n
        else:
            n_train = int(n * args.train_ratio)
            n_val = n - n_train
            n_test = 0
            splits["train"].extend(pairs[:n_train])
            splits["val"].extend(pairs[n_train:])

        shape_stats[shape_id] = {
            "num_sizes": len(sizes),
            "num_pairs": n,
            "train": n_train,
            "val": n_val,
            "test": n_test,
        }

    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    payload = {
        "created_at": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
        "seed": args.seed,
        "mode": "old_xlsx_split",
        "old_xlsx": str(old_xlsx_path),
        "train_ratio": args.train_ratio,
        "val_ratio": args.val_ratio,
        "exclude_self": bool(args.exclude_self),
        "test_shapes": len(test_shapes),
        "shape_index": str(shape_index_path),
        "size_scale_factors": str(size_scale_path),
        "elements_embed": str(elements_path),
        "shape_element_map": str(shape_elements_path),
        "filter_missing_elements": bool(valid_layouts is not None),
        "filter_pdf_pairs": bool(shape_pdf_map is not None),
        "filtered_jsons": dropped_jsons,
        "filtered_shapes": dropped_shapes,
        "filtered_shape_reasons": dropped_shape_reasons,
        "filtered_pairs": dropped_pairs,
        "shapes": shape_stats,
        "splits": splits,
    }
    out_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"shapes: {len(shapes)}")
    print(f"test shapes: {len(test_shapes)}")
    print(f"dropped shapes: {dropped_shapes} reasons={dropped_shape_reasons}")
    print(f"dropped jsons: {dropped_jsons}")
    print(f"dropped pairs (pdf filter): {dropped_pairs}")
    print(f"train pairs: {len(splits['train'])}")
    print(f"val pairs: {len(splits['val'])}")
    print(f"test pairs: {len(splits['test'])}")
    print(f"output: {out_path}")


if __name__ == "__main__":
    main()
