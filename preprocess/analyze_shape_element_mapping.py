#!/usr/bin/env python3
import argparse
import json
import sys
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path

try:
    import openpyxl
    from openpyxl.utils import column_index_from_string
except ImportError as exc:
    raise SystemExit("openpyxl is required to run this script") from exc

COL_ID = 1      # A
COL_SAMPLE = 2  # B (样品号)
COL_JSON = 5    # E
COL_PDF = 6     # F
COL_SHAPE = 4   # D
COL_X = 11      # K
COL_Y = 12      # L
COL_W = 13      # M
COL_H = 14      # N
DEFAULT_TYPE_COL = 8  # H


def parse_column(value):
    if value is None:
        return None
    if isinstance(value, int):
        return value
    text = str(value).strip()
    if text == "":
        return None
    if text.isdigit():
        return int(text)
    try:
        return column_index_from_string(text.upper())
    except Exception:
        return None


def to_float(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        value = value.strip()
        if value == "":
            return None
        try:
            return float(value)
        except ValueError:
            return None
    return None


def _make_rel_id(path: Path, project_root: Path, fallback_root: Path | None = None) -> str:
    try:
        return path.resolve().relative_to(project_root).as_posix()
    except Exception:
        if fallback_root is not None:
            try:
                return path.resolve().relative_to(fallback_root).as_posix()
            except Exception:
                pass
        return path.resolve().as_posix()


def resolve_path(root: Path, value) -> Path:
    rel = Path(str(value))
    if rel.is_absolute():
        return rel
    return root / rel


def extract_style_size_piece(json_rel: str):
    parts = Path(json_rel).parts
    if len(parts) < 3:
        return None, None, Path(json_rel).stem
    style = parts[-3]
    size = parts[-2]
    piece = Path(parts[-1]).stem
    return style, size, piece


def main():
    parser = argparse.ArgumentParser(
        description="Group pattern pieces by sample/style/shape name (columns B/D) and report conflicts."
    )
    parser.add_argument("xlsx", help="Path to input xlsx file.")
    parser.add_argument(
        "--root",
        help="Root directory for relative JSON paths (defaults to project_root/pattern/pattern if exists).",
    )
    parser.add_argument(
        "--type-col",
        help="Type column (letter or 1-based index). Default: H. Used only for stats; not required.",
    )
    parser.add_argument(
        "--shape-col",
        help="Shape name column (letter or 1-based index). Default: D.",
    )
    parser.add_argument(
        "--sample-col",
        help="Sample id column (letter or 1-based index). Default: B.",
    )
    parser.add_argument(
        "--out",
        help="Output json path (default: <project>/data/shape_element_map.json).",
    )
    parser.add_argument(
        "--out-index",
        help="Output shape-size index json (default: <project>/data/shape_size_index.json).",
    )
    parser.add_argument(
        "--out-json-lookup",
        help="Output json->size map (default: <project>/data/json_size_lookup.json).",
    )
    parser.add_argument(
        "--out-pdf-map",
        help="Output pdf->shape map (default: <project>/data/pdf_shape_map.json).",
    )
    parser.add_argument(
        "--out-conflicts",
        help="Output conflict report path (default: <project>/data/shape_conflict_report.txt).",
    )
    args = parser.parse_args()

    project_root = Path(__file__).resolve().parents[1]
    default_root = project_root / "pattern" / "pattern"
    if not default_root.is_dir():
        default_root = project_root
    root = Path(args.root) if args.root else default_root

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.is_file():
        print(f"Not a file: {xlsx_path}", file=sys.stderr)
        return 1

    out_path = Path(args.out) if args.out else project_root / "data" / "shape_element_map.json"
    out_path.parent.mkdir(parents=True, exist_ok=True)

    out_index = Path(args.out_index) if args.out_index else project_root / "data" / "shape_size_index.json"
    out_index.parent.mkdir(parents=True, exist_ok=True)

    out_json_lookup = Path(args.out_json_lookup) if args.out_json_lookup else project_root / "data" / "json_size_lookup.json"
    out_json_lookup.parent.mkdir(parents=True, exist_ok=True)

    out_pdf_map = Path(args.out_pdf_map) if args.out_pdf_map else project_root / "data" / "pdf_shape_map.json"
    out_pdf_map.parent.mkdir(parents=True, exist_ok=True)

    out_conflicts = Path(args.out_conflicts) if args.out_conflicts else project_root / "data" / "shape_conflict_report.txt"
    out_conflicts.parent.mkdir(parents=True, exist_ok=True)

    type_col = parse_column(args.type_col) if args.type_col else DEFAULT_TYPE_COL
    shape_col = parse_column(args.shape_col) if args.shape_col else COL_SHAPE
    sample_col = parse_column(args.sample_col) if args.sample_col else COL_SAMPLE

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    pattern_map = {}
    shape_groups = {}
    pdf_to_shapes = defaultdict(set)
    size_conflicts = []
    skipped = 0

    for row in range(2, ws.max_row + 1):
        element_id = ws.cell(row=row, column=COL_ID).value
        sample_cell = ws.cell(row=row, column=sample_col).value if sample_col else None
        json_cell = ws.cell(row=row, column=COL_JSON).value
        pdf_cell = ws.cell(row=row, column=COL_PDF).value
        shape_cell = ws.cell(row=row, column=shape_col).value
        type_cell = ws.cell(row=row, column=type_col).value if type_col else None

        if element_id is None or str(element_id).strip() == "":
            continue
        if sample_cell is None or str(sample_cell).strip() == "":
            skipped += 1
            continue
        if json_cell is None or str(json_cell).strip() == "":
            skipped += 1
            continue
        if pdf_cell is None or str(pdf_cell).strip() == "":
            skipped += 1
            continue
        if shape_cell is None or str(shape_cell).strip() == "":
            skipped += 1
            continue
        if type_cell is None or str(type_cell).strip() == "":
            skipped += 1
            continue
        x = to_float(ws.cell(row=row, column=COL_X).value)
        y = to_float(ws.cell(row=row, column=COL_Y).value)
        w = to_float(ws.cell(row=row, column=COL_W).value)
        h = to_float(ws.cell(row=row, column=COL_H).value)
        if None in (x, y, w, h):
            skipped += 1
            continue

        sample_id = str(sample_cell).strip()
        element_id_str = str(element_id).strip()
        element_uid = f"{sample_id}::{element_id_str}"
        shape_name_raw = str(shape_cell).strip()

        json_path = resolve_path(root, json_cell)
        json_rel = _make_rel_id(json_path, project_root, fallback_root=root)

        pdf_path = resolve_path(project_root, pdf_cell)
        pdf_rel = _make_rel_id(pdf_path, project_root, fallback_root=project_root)

        type_name = str(type_cell).strip()

        style, size, piece = extract_style_size_piece(json_rel)
        if style:
            shape_name = f"{sample_id}-{style}-{shape_name_raw}"
        else:
            shape_name = f"{sample_id}-{shape_name_raw}"

        entry = pattern_map.setdefault(json_rel, {
            "json_path": str(json_path),
            "pdf_paths": set(),
            "sample_id": sample_id,
            "style": style,
            "size": size,
            "piece": piece,
            "shape_name": shape_name,
            "element_count": 0,
            "type_counts": Counter(),
        })
        entry["pdf_paths"].add(pdf_rel)
        entry["element_count"] += 1
        if type_name is not None:
            entry["type_counts"][type_name] += 1

        group = shape_groups.setdefault(shape_name, {
            "sizes": {},
            "pdfs": set(),
            "jsons": set(),
            "size_elements": {},
            "json_elements": {},
        })
        group["pdfs"].add(pdf_rel)
        group["jsons"].add(json_rel)
        if size:
            elem_set = group["size_elements"].setdefault(size, set())
            elem_set.add(element_uid)
        json_entry = group["json_elements"].setdefault(
            json_rel, {"pdfs": set(), "elements": set()}
        )
        if pdf_rel:
            json_entry["pdfs"].add(pdf_rel)
        json_entry["elements"].add(element_uid)

        if size:
            if size in group["sizes"] and group["sizes"][size] != json_rel:
                print(
                    "[size conflict] "
                    f"shape={shape_name} size={size} existing={group['sizes'][size]} "
                    f"incoming={json_rel} pdf={pdf_rel} row={row}",
                    file=sys.stderr,
                )
                size_conflicts.append({
                    "shape_name": shape_name,
                    "size": size,
                    "existing": group["sizes"][size],
                    "incoming": json_rel,
                    "incoming_pdf": pdf_rel,
                })
            else:
                group["sizes"][size] = json_rel

        if pdf_rel:
            pdf_to_shapes[pdf_rel].add(shape_name)

    # Build shapes using grouped shape name (sample/style/shape)
    shapes = {}
    shape_size_index = {}
    pdf_shape_map = {}
    max_layout_element_count = 0

    for shape_name, group in shape_groups.items():
        size_map = dict(group["sizes"])
        size_elements = {
            size: sorted(ids) for size, ids in group["size_elements"].items()
        }
        size_element_counts = {
            size: len(ids) for size, ids in size_elements.items()
        }
        json_elements = {
            json_rel: {
                "pdfs": sorted(info["pdfs"]),
                "elements": sorted(info["elements"]),
            }
            for json_rel, info in group["json_elements"].items()
        }
        shapes[shape_name] = {
            "pdfs": sorted(group["pdfs"]),
            "sizes": size_map,
            "jsons": sorted(group["jsons"]),
            "size_elements": size_elements,
            "size_element_counts": size_element_counts,
            "json_elements": json_elements,
        }
        shape_max = max(size_element_counts.values(), default=0)
        if shape_max > max_layout_element_count:
            max_layout_element_count = shape_max
        shape_size_index[shape_name] = size_map

    for pdf, shape_set in pdf_to_shapes.items():
        shape_list = sorted(shape_set)
        if shape_list:
            pdf_shape_map[pdf] = shape_list

    # Build json_size_lookup (shape_id -> json -> size_map)
    json_size_lookup = {}
    for shape_name, group in shape_groups.items():
        size_map = shape_size_index.get(shape_name, {})
        shape_lookup = json_size_lookup.setdefault(shape_name, {})
        for json_rel in group["jsons"]:
            shape_lookup[json_rel] = size_map

    out_index.write_text(json.dumps(shape_size_index, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    out_json_lookup.write_text(json.dumps(json_size_lookup, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    out_pdf_map.write_text(json.dumps(pdf_shape_map, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    if size_conflicts:
        lines = []
        if size_conflicts:
            lines.append("[size->json conflicts]")
            for conflict in size_conflicts:
                lines.append(f"- shape_name: {conflict['shape_name']}")
                lines.append(f"  size: {conflict['size']}")
                lines.append(f"  existing: {conflict['existing']}")
                lines.append(f"  incoming: {conflict['incoming']}")
                if conflict.get("incoming_pdf"):
                    lines.append(f"  incoming_pdf: {conflict['incoming_pdf']}")
        out_conflicts.write_text("\n".join(lines) + "\n", encoding="utf-8")
    else:
        out_conflicts.write_text("(no conflicts)\n", encoding="utf-8")

    result = {
        "created_at": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
        "project_root": str(project_root),
        "json_root": str(root),
        "xlsx": str(xlsx_path),
        "type_col": int(type_col) if type_col else None,
        "sample_col": int(sample_col) if sample_col else None,
        "skipped_rows": skipped,
        "max_layout_element_count": max_layout_element_count,
        "size_conflicts": size_conflicts,
        "shape_size_index_path": str(out_index),
        "json_size_lookup_path": str(out_json_lookup),
        "pdf_shape_map_path": str(out_pdf_map),
        "conflict_report_path": str(out_conflicts),
        "shapes": shapes,
    }

    out_path.write_text(json.dumps(result, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    print(f"pdfs: {len(pdf_shape_map)}")
    print(f"shapes: {len(shapes)}")
    print(f"size_conflicts: {len(size_conflicts)}")
    print(f"shape_size_index: {out_index}")
    print(f"json_size_lookup: {out_json_lookup}")
    print(f"pdf_shape_map: {out_pdf_map}")
    print(f"conflict_report: {out_conflicts}")
    print(f"output: {out_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
