#!/usr/bin/env python3
import argparse
import json
import sys
from datetime import datetime, timezone
from pathlib import Path

try:
    import numpy as np
except ImportError as exc:
    raise SystemExit("numpy is required to run this script") from exc

try:
    import openpyxl
    from openpyxl.utils import column_index_from_string
except ImportError as exc:
    raise SystemExit("openpyxl is required to run this script") from exc

try:
    import torch
except ImportError as exc:
    raise SystemExit("torch is required to run this script") from exc

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from normalize_xlsx import process_sheet, resolve_json_path, COL_JSON, COL_X, COL_Y, COL_W, COL_H, COL_SCALE, NORM_RANGE
from ly import FourierFeatureEncoder

COL_ID = 1
COL_PDF = 6
COL_LOGO = 9


def to_float(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        value = value.strip()
        if value == "":
            return None
        try:
            return float(value)
        except ValueError:
            return None
    return None




def to_int(value):
    if value is None:
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    if isinstance(value, str):
        value = value.strip()
        if value == "":
            return None
        try:
            return int(float(value))
        except ValueError:
            return None
    return None

def _make_rel_id(path: Path, project_root: Path, fallback_root: Path | None = None) -> str:
    # Prefer paths relative to LayoutFormer repo root.
    # If the file is outside project_root, fall back to fallback_root (scan root),
    # otherwise return an absolute posix path.
    try:
        return path.resolve().relative_to(project_root).as_posix()
    except Exception:
        if fallback_root is not None:
            try:
                return path.resolve().relative_to(fallback_root).as_posix()
            except Exception:
                pass
        return path.resolve().as_posix()


def parse_column(value):
    if value is None:
        return None
    if isinstance(value, int):
        return value
    text = str(value).strip()
    if text == "":
        return None
    if text.isdigit():
        return int(text)
    try:
        return column_index_from_string(text.upper())
    except Exception:
        return None


def detect_type_col(ws):
    header_map = {}
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value is None:
            continue
        key = str(cell.value).strip().lower()
        if key:
            header_map[key] = idx
    for key in ("type", "label", "class", "??", "??", "??"):
        if key in header_map:
            return header_map[key]
    return None




def _safe_out_path(path: Path) -> Path:
    if not path.exists():
        return path
    try:
        path.unlink()
        return path
    except PermissionError:
        stem = path.stem
        suffix = path.suffix
        for i in range(1, 1000):
            cand = path.with_name(f"{stem}-{i}{suffix}")
            if not cand.exists():
                return cand
        raise


def normalize_xlsx(xlsx_path: Path, root: Path):
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    process_sheet(ws, root)
    out_path = xlsx_path.with_name(xlsx_path.stem + "-n" + xlsx_path.suffix)
    out_path = _safe_out_path(out_path)
    wb.save(out_path)
    return out_path


def load_vocab(path: Path):
    if not path.exists():
        return None
    data = json.loads(path.read_text(encoding="utf-8"))
    type_to_id = data.get("type_to_id")
    if not isinstance(type_to_id, dict):
        return None
    id_to_type = {int(v): k for k, v in type_to_id.items()}
    return {
        "type_to_id": {k: int(v) for k, v in type_to_id.items()},
        "id_to_type": id_to_type,
    }


def save_vocab(path: Path, type_to_id: dict):
    id_to_type = {v: k for k, v in type_to_id.items()}
    payload = {
        "type_to_id": type_to_id,
        "id_to_type": id_to_type,
        "num_types": len(type_to_id),
        "created_at": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
    }
    path.write_text(json.dumps(payload, ensure_ascii=True, indent=2) + "\n", encoding="utf-8")


def main():
    parser = argparse.ArgumentParser(
        description="Normalize an xlsx, build element embeddings, and save type vocab + embeddings."
    )
    parser.add_argument("xlsx", help="Path to input xlsx file.")
    parser.add_argument(
        "--root",
        help="Root directory for relative JSON paths (defaults to project_root/pattern/pattern if exists).",
    )
    parser.add_argument(
        "--pdf-root",
        help="Root directory for relative PDF paths (defaults to project_root).",
    )
    parser.add_argument(
        "--type-col",
        help="Type column (letter or 1-based index). If omitted, will try header names.",
    )
    parser.add_argument("--num-bands", type=int, default=10, help="Fourier bands (L).")
    parser.add_argument("--max-freq", type=float, default=10.0, help="Max freq (kept for API).")
    parser.add_argument(
        "--out-dir",
        help="Output directory (default: <project>/data/element_emb).",
    )
    parser.add_argument(
        "--vocab",
        help="Type vocab json (default: <out-dir>/type_vocab.json). If exists, will be reused and extended.",
    )
    args = parser.parse_args()

    project_root = ROOT
    default_root = project_root / "pattern" / "pattern"
    if not default_root.is_dir():
        default_root = project_root

    root = Path(args.root) if args.root else default_root
    pdf_root = Path(args.pdf_root) if args.pdf_root else project_root

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.is_file():
        print(f"Not a file: {xlsx_path}", file=sys.stderr)
        return 1

    out_dir = Path(args.out_dir) if args.out_dir else project_root / "data" / "element_emb"
    out_dir.mkdir(parents=True, exist_ok=True)

    vocab_path = Path(args.vocab) if args.vocab else out_dir / "type_vocab.json"

    # Step 1: normalize -> -n.xlsx
    norm_path = normalize_xlsx(xlsx_path, root)

    # Step 2: read normalized data
    wb = openpyxl.load_workbook(norm_path)
    ws = wb.active

    type_col = parse_column(args.type_col) if args.type_col else None
    if type_col is None:
        type_col = detect_type_col(ws)
    if type_col is None:
        type_col = column_index_from_string("H")

    rows = []
    types = set()
    skipped = 0

    for row in range(2, ws.max_row + 1):
        element_id = ws.cell(row=row, column=COL_ID).value
        if element_id is None or str(element_id).strip() == "":
            continue

        json_cell = ws.cell(row=row, column=COL_JSON).value
        pdf_cell = ws.cell(row=row, column=COL_PDF).value
        type_cell = ws.cell(row=row, column=type_col).value
        logo_cell = ws.cell(row=row, column=COL_LOGO).value

        if json_cell is None or str(json_cell).strip() == "":
            skipped += 1
            continue
        if type_cell is None or str(type_cell).strip() == "":
            skipped += 1
            continue

        logo_level = to_int(logo_cell)
        x = to_float(ws.cell(row=row, column=COL_X).value)
        y = to_float(ws.cell(row=row, column=COL_Y).value)
        w = to_float(ws.cell(row=row, column=COL_W).value)
        h = to_float(ws.cell(row=row, column=COL_H).value)
        if None in (x, y, w, h):
            skipped += 1
            continue

        # Normalize paths to project_root relative
        json_path = resolve_json_path(root, json_cell)
        json_rel = _make_rel_id(json_path, project_root, fallback_root=root)

        pdf_rel = ""
        if pdf_cell is not None and str(pdf_cell).strip() != "":
            pdf_path = resolve_json_path(pdf_root, pdf_cell)
            pdf_rel = _make_rel_id(pdf_path, project_root, fallback_root=pdf_root)

        type_name = str(type_cell).strip()
        types.add(type_name)

        rows.append({
            "element_id": str(element_id).strip(),
            "pdf_path": pdf_rel,
            "pattern_json": json_rel,
            "type_name": type_name,
            "center_x": float(x),
            "center_y": float(y),
            "w": float(w),
            "h": float(h),
            "logo_level": logo_level,
        })

    if not rows:
        print("No valid rows found.", file=sys.stderr)
        return 3

    # Step 3: build or extend vocab
    vocab = load_vocab(vocab_path)
    if vocab is None:
        type_to_id = {t: i for i, t in enumerate(sorted(types))}
    else:
        type_to_id = vocab["type_to_id"]
        next_id = max(type_to_id.values(), default=-1) + 1
        new_types = sorted(t for t in types if t not in type_to_id)
        for t in new_types:
            type_to_id[t] = next_id
            next_id += 1
    save_vocab(vocab_path, type_to_id)

    # Step 4: labels and embeddings
    labels = np.array([type_to_id[r["type_name"]] for r in rows], dtype=np.int64)
    center_x = np.array([r["center_x"] for r in rows], dtype=np.float32)
    center_y = np.array([r["center_y"] for r in rows], dtype=np.float32)
    widths = np.array([r["w"] for r in rows], dtype=np.float32)
    heights = np.array([r["h"] for r in rows], dtype=np.float32)
    logo_levels = np.array([r["logo_level"] if r.get("logo_level") is not None else -1 for r in rows], dtype=np.int64)
    sin = np.zeros(len(rows), dtype=np.float32)
    cos = np.ones(len(rows), dtype=np.float32)

    encoder = FourierFeatureEncoder(num_bands=args.num_bands, max_freq=args.max_freq)
    with torch.no_grad():
        x_tensor = torch.from_numpy(center_x).unsqueeze(0)
        y_tensor = torch.from_numpy(center_y).unsqueeze(0)
        fx = encoder(x_tensor)
        fy = encoder(y_tensor)
        emb = torch.cat([fx, fy], dim=-1).squeeze(0).cpu().numpy().astype(np.float32)

    # Step 5: save
    out_npz = out_dir / "elements_embed.npz"
    out_meta = out_dir / "elements_meta.json"

    element_ids = np.array([r["element_id"] for r in rows], dtype=object)
    pdf_paths = np.array([r["pdf_path"] for r in rows], dtype=object)
    pattern_jsons = np.array([r["pattern_json"] for r in rows], dtype=object)

    np.savez(
        out_npz,
        element_ids=element_ids,
        pdf_paths=pdf_paths,
        pattern_json_paths=pattern_jsons,
        labels=labels,
        center_x=center_x,
        center_y=center_y,
        w=widths,
        h=heights,
        logo_level=logo_levels,
        sin=sin,
        cos=cos,
        element_embed=emb,
    )

    element_id_to_index = {rows[i]["element_id"]: i for i in range(len(rows))}

    meta = {
        "count": len(rows),
        "skipped": skipped,
        "num_bands": args.num_bands,
        "embed_dim": int(emb.shape[1]),
        "coord_origin": "bottom-left",
        "paths_relative_to": "project_root",
        "project_root": str(project_root),
        "json_root": str(root),
        "pdf_root": str(pdf_root),
        "xlsx_source": str(xlsx_path),
        "xlsx_norm": str(norm_path),
        "norm_range": NORM_RANGE,
        "type_col": int(type_col),
        "columns": {
            "element_id": COL_ID,
            "pattern_json": COL_JSON,
            "pdf_path": COL_PDF,
            "center_x": COL_X,
            "center_y": COL_Y,
            "w": COL_W,
            "h": COL_H,
            "logo_level": COL_LOGO,
            "scale": COL_SCALE,
        },
        "npz_keys": {
            "element_ids": "element_ids",
            "pdf_paths": "pdf_paths",
            "pattern_json_paths": "pattern_json_paths",
            "labels": "labels",
            "center_x": "center_x",
            "center_y": "center_y",
            "w": "w",
            "h": "h",
            "logo_level": "logo_level",
            "sin": "sin",
            "cos": "cos",
            "element_embed": "element_embed",
        },
        "type_vocab": str(vocab_path),
        "element_id_to_index": element_id_to_index,
        "created_at": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
    }
    out_meta.write_text(json.dumps(meta, ensure_ascii=True, indent=2) + "\n", encoding="utf-8")

    print(f"normalized: {norm_path}")
    print(f"elements: {len(rows)} (skipped {skipped})")
    print(f"npz: {out_npz}")
    print(f"meta: {out_meta}")
    print(f"vocab: {vocab_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
